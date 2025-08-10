#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ENHANCED SCRAPER RUNNER
Advanced scrapers to reach 10,000 unique phones for each category
"""

import subprocess
import time
import os
import sys
import logging
from pathlib import Path

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('enhanced_runner.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def check_dependencies():
    """Check if all required packages are installed"""
    logger.info("🔍 Checking dependencies...")
    
    required_packages = [
        ('requests', 'requests'),
        ('bs4', 'beautifulsoup4'),
        ('openpyxl', 'openpyxl'),
        ('lxml', 'lxml')
    ]
    
    missing_packages = []
    
    for import_name, package_name in required_packages:
        try:
            __import__(import_name)
            logger.info(f"✅ {package_name} is installed")
        except ImportError:
            missing_packages.append(package_name)
            logger.error(f"❌ {package_name} is missing")
    
    if missing_packages:
        logger.error(f"Missing packages: {', '.join(missing_packages)}")
        logger.info("Please install missing packages with: pip install " + " ".join(missing_packages))
        return False
    
    logger.info("✅ All dependencies are installed!")
    return True

def run_enhanced_agent_scraper():
    """Run the enhanced agent scraper"""
    logger.info("🚀 Starting ENHANCED AGENT SCRAPER...")
    
    try:
        # Run the enhanced agent scraper
        result = subprocess.run([
            sys.executable, "mega_agent_scraper_enhanced.py"
        ], capture_output=True, text=True, encoding='utf-8')
        
        if result.returncode == 0:
            logger.info("✅ Enhanced agent scraper completed successfully!")
            logger.info(f"Output: {result.stdout}")
        else:
            logger.error(f"❌ Enhanced agent scraper failed with return code {result.returncode}")
            logger.error(f"Error: {result.stderr}")
            return False
            
    except Exception as e:
        logger.error(f"❌ Error running enhanced agent scraper: {e}")
        return False
    
    return True

def run_enhanced_owner_scraper():
    """Run the enhanced owner scraper"""
    logger.info("🚀 Starting ENHANCED OWNER SCRAPER...")
    
    try:
        # Run the enhanced owner scraper
        result = subprocess.run([
            sys.executable, "turbo_owner_scraper_enhanced.py"
        ], capture_output=True, text=True, encoding='utf-8')
        
        if result.returncode == 0:
            logger.info("✅ Enhanced owner scraper completed successfully!")
            logger.info(f"Output: {result.stdout}")
        else:
            logger.error(f"❌ Enhanced owner scraper failed with return code {result.returncode}")
            logger.error(f"Error: {result.stderr}")
            return False
            
    except Exception as e:
        logger.error(f"❌ Error running enhanced owner scraper: {e}")
        return False
    
    return True

def check_results():
    """Check the results of both scrapers"""
    logger.info("📊 Checking results...")
    
    results = {}
    
    # Check agent results
    if os.path.exists('agents.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook('agents.xlsx')
            ws = wb.active
            agent_count = ws.max_row - 1  # Subtract header row
            results['agents'] = agent_count
            logger.info(f"📱 Agent phones collected: {agent_count}")
        except Exception as e:
            logger.error(f"Error reading agents.xlsx: {e}")
            results['agents'] = 0
    else:
        logger.warning("agents.xlsx not found")
        results['agents'] = 0
    
    # Check owner results
    if os.path.exists('owners.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook('owners.xlsx')
            ws = wb.active
            owner_count = ws.max_row - 1  # Subtract header row
            results['owners'] = owner_count
            logger.info(f"🏠 Owner phones collected: {owner_count}")
        except Exception as e:
            logger.error(f"Error reading owners.xlsx: {e}")
            results['owners'] = 0
    else:
        logger.warning("owners.xlsx not found")
        results['owners'] = 0
    
    total = results.get('agents', 0) + results.get('owners', 0)
    logger.info(f"🎯 TOTAL phones collected: {total}")
    
    # Check targets
    agent_target = 10000
    owner_target = 10000
    total_target = 20000
    
    agent_success = results.get('agents', 0) >= agent_target
    owner_success = results.get('owners', 0) >= owner_target
    total_success = total >= total_target
    
    logger.info("📋 TARGET ACHIEVEMENT:")
    logger.info(f"   Agents: {results.get('agents', 0)}/{agent_target} {'✅' if agent_success else '❌'}")
    logger.info(f"   Owners: {results.get('owners', 0)}/{owner_target} {'✅' if owner_success else '❌'}")
    logger.info(f"   Total:  {total}/{total_target} {'✅' if total_success else '❌'}")
    
    return results

def main():
    """Main execution"""
    logger.info("🚀 ENHANCED SCRAPER RUNNER STARTING")
    logger.info("🎯 Target: 10,000 unique agent phones + 10,000 unique owner phones")
    logger.info("=" * 60)
    
    # Check dependencies
    if not check_dependencies():
        logger.error("❌ Dependencies check failed. Exiting.")
        return
    
    # Run enhanced agent scraper
    logger.info("\n" + "=" * 60)
    if not run_enhanced_agent_scraper():
        logger.error("❌ Enhanced agent scraper failed. Exiting.")
        return
    
    # Run enhanced owner scraper
    logger.info("\n" + "=" * 60)
    if not run_enhanced_owner_scraper():
        logger.error("❌ Enhanced owner scraper failed. Exiting.")
        return
    
    # Check final results
    logger.info("\n" + "=" * 60)
    results = check_results()
    
    # Final summary
    logger.info("\n" + "=" * 60)
    logger.info("🏁 ENHANCED SCRAPING COMPLETE!")
    
    total = results.get('agents', 0) + results.get('owners', 0)
    if total >= 20000:
        logger.info("🎉 SUCCESS: Target of 20,000 total unique phones achieved!")
    else:
        logger.info(f"📊 PARTIAL SUCCESS: Collected {total} phones (target was 20,000)")
    
    logger.info("📁 Output files:")
    logger.info("   - agents.xlsx: Agent phone numbers")
    logger.info("   - owners.xlsx: Owner phone numbers")
    logger.info("   - enhanced_runner.log: Detailed execution log")

if __name__ == "__main__":
    main()
