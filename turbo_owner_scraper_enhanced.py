#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ENHANCED Turbo Owner Phone Scraper
Advanced discovery methods to reach 10,000+ unique phones
"""

import csv
import logging
import time
import re
import json
from typing import Set, List, Optional
from urllib.parse import urljoin, urlparse
import threading
from scraper.fetcher import MyHomeFetcher

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('enhanced_owner_scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EnhancedOwnerScraper:
    """Enhanced owner scraper with multiple discovery methods"""
    
    def __init__(self):
        self.fetcher = MyHomeFetcher()
        self.owner_phones = set()
        self.processed_urls = set()
        self.target_phones = 10000
        self.discovered_urls = set()
        
    def load_existing_phones(self):
        """Load existing phones to avoid duplicates"""
        existing_files = [
            'agents.csv',
            'owners_direct.csv', 
            'owners.csv',
            'master_phones_only.csv'
        ]
        
        for filename in existing_files:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        phone = row.get('Phone', '').strip()
                        if phone:
                            normalized = self.normalize_phone(phone)
                            if normalized:
                                self.owner_phones.add(normalized)
                logger.info(f"Loaded {len(self.owner_phones)} existing phones from {filename}")
            except Exception as e:
                logger.debug(f"Could not load {filename}: {e}")
        
        logger.info(f"Starting with {len(self.owner_phones)} existing unique phones")
    
    def normalize_phone(self, phone: str) -> str:
        """Normalize phone for deduplication"""
        if not phone:
            return ""
        
        digits = re.sub(r'\D', '', phone)
        
        if len(digits) == 9 and digits.startswith('5'):
            return f"+995{digits}"
        elif len(digits) == 12 and digits.startswith('995'):
            return f"+{digits}"
        elif len(digits) >= 9:
            return f"+995{digits[-9:]}"
        
        return ""
    
    def discover_property_urls_comprehensive(self, max_urls: int = 50000) -> List[str]:
        """Comprehensive property URL discovery using multiple methods"""
        logger.info(f"🔍 COMPREHENSIVE property URL discovery (target: {max_urls})")
        
        all_urls = set()
        
        # Method 1: Multiple API endpoints with extensive pagination
        api_endpoints = [
            "/statements",
            "/statements?operation_type_id=1",  # Rent
            "/statements?operation_type_id=3",  # Sale
            "/statements?operation_type_id=2",  # Daily rent
            "/statements?operation_type_id=4",  # Commercial
        ]
        
        for endpoint in api_endpoints:
            logger.info(f"📡 Fetching from {endpoint}")
            
            for page in range(1, 1000):  # Much more extensive pagination
                try:
                    api_data = self.fetcher.fetch_property_listings_api(page, endpoint)
                    if not api_data or not api_data.get('result'):
                        logger.info(f"No more data from {endpoint} at page {page}")
                        break
                    
                    data = api_data.get('data', {})
                    if isinstance(data, dict):
                        items = data.get('data', [])
                    else:
                        items = data
                    
                    if not items:
                        logger.info(f"Empty page {page} for {endpoint}")
                        break
                    
                    page_urls = []
                    for item in items:
                        if isinstance(item, dict) and 'statement_id' in item:
                            url = f"https://www.myhome.ge/pr/{item['statement_id']}/"
                            page_urls.append(url)
                    
                    all_urls.update(page_urls)
                    logger.info(f"{endpoint} page {page}: {len(page_urls)} URLs. Total unique: {len(all_urls)}")
                    
                    if len(all_urls) >= max_urls:
                        logger.info(f"✅ Reached target of {max_urls} URLs!")
                        return list(all_urls)[:max_urls]
                        
                    time.sleep(0.1)  # Fast rate limiting
                    
                except Exception as e:
                    logger.error(f"Error fetching {endpoint} page {page}: {e}")
                    break
        
        # Method 2: Discover through agent properties
        logger.info("🔍 Discovering properties through agents...")
        try:
            agent_urls = self.discover_properties_through_agents(10000)
            all_urls.update(agent_urls)
            logger.info(f"Added {len(agent_urls)} URLs through agents. Total: {len(all_urls)}")
        except Exception as e:
            logger.error(f"Error discovering through agents: {e}")
        
        # Method 3: Discover through property categories
        logger.info("🔍 Discovering through property categories...")
        try:
            category_urls = self.discover_through_categories(10000)
            all_urls.update(category_urls)
            logger.info(f"Added {len(category_urls)} URLs through categories. Total: {len(all_urls)}")
        except Exception as e:
            logger.error(f"Error discovering through categories: {e}")
        
        logger.info(f"✅ COMPREHENSIVE discovery complete: {len(all_urls)} unique URLs")
        return list(all_urls)[:max_urls]
    
    def discover_properties_through_agents(self, max_urls: int = 10000) -> Set[str]:
        """Discover properties by going through agent listings"""
        agent_urls = set()
        
        # Get agents first
        for page in range(1, 200):  # Many agent pages
            try:
                api_data = self.fetcher.fetch_agents_api(page)
                if not api_data or not api_data.get('result'):
                    break
                
                agents = api_data.get('data', {}).get('data', [])
                if not agents:
                    break
                
                for agent in agents:
                    if 'id' in agent:
                        agent_id = agent['id']
                        # Get agent's properties
                        try:
                            agent_props = self.fetcher.fetch_agent_sub_agents_api(agent_id, 1)
                            if agent_props and agent_props.get('result'):
                                props_data = agent_props.get('data', {}).get('data', [])
                                for prop in props_data:
                                    if 'statement_id' in prop:
                                        url = f"https://www.myhome.ge/pr/{prop['statement_id']}/"
                                        agent_urls.add(url)
                                        
                                        if len(agent_urls) >= max_urls:
                                            return agent_urls
                        except Exception as e:
                            logger.debug(f"Error getting agent {agent_id} properties: {e}")
                            continue
                
                time.sleep(0.2)
                
            except Exception as e:
                logger.error(f"Error fetching agents page {page}: {e}")
                break
        
        return agent_urls
    
    def discover_through_categories(self, max_urls: int = 10000) -> Set[str]:
        """Discover through property category pages"""
        category_urls = set()
        
        # Property categories
        categories = [
            "rent", "sale", "daily-rent", "commercial",
            "apartment", "house", "land", "office"
        ]
        
        for category in categories:
            try:
                # Try to get category listings
                for page in range(1, 100):
                    try:
                        # Construct category URL
                        category_url = f"https://www.myhome.ge/{category}/?page={page}"
                        response = self.fetcher.fetch_page(category_url)
                        
                        if not response:
                            break
                        
                        # Extract property URLs from HTML
                        urls = re.findall(r'href="(/pr/\d+/[^"]*)"', response)
                        for url in urls:
                            full_url = f"https://www.myhome.ge{url}"
                            category_urls.add(full_url)
                            
                            if len(category_urls) >= max_urls:
                                return category_urls
                        
                        if len(urls) == 0:
                            break
                            
                        time.sleep(0.1)
                        
                    except Exception as e:
                        logger.debug(f"Error in category {category} page {page}: {e}")
                        break
                        
            except Exception as e:
                logger.error(f"Error processing category {category}: {e}")
                continue
        
        return category_urls
    
    def scrape_property_phone_enhanced(self, property_url: str) -> Optional[str]:
        """Enhanced phone scraping with multiple methods"""
        try:
            # Method 1: Direct page fetch
            response = self.fetcher.fetch_page(property_url)
            if response:
                phone = self.extract_phone_from_html(response)
                if phone:
                    return phone
            
            # Method 2: Try to extract statement UUID and use API
            try:
                uuid = self.extract_statement_uuid_from_url(property_url)
                if uuid:
                    api_response = self.fetcher.fetch_property_phone_api(uuid)
                    if api_response and api_response.get('result'):
                        phone_data = api_response.get('data', {})
                        if 'phone_number' in phone_data:
                            phone = phone_data['phone_number']
                            normalized = self.normalize_phone(phone)
                            if normalized:
                                return normalized
            except Exception as e:
                logger.debug(f"API method failed for {property_url}: {e}")
            
            return None
            
        except Exception as e:
            logger.debug(f"Error scraping {property_url}: {e}")
            return None
    
    def extract_phone_from_html(self, html_content: str) -> Optional[str]:
        """Extract phone from HTML content"""
        # Enhanced phone patterns
        phone_patterns = [
            r'\+995\s*\d{3}\s*\d{3}\s*\d{3}',
            r'995\d{9}',
            r'\b5\d{8}\b',
            r'"phone_number":\s*"(\d+)"',
            r'tel:(\d+)',
            r'(\+995\s*\d{3}\s*\d{3}\s*\d{3})',
            r'(995\d{9})',
            r'(\b5\d{8}\b)',
        ]
        
        for pattern in phone_patterns:
            matches = re.findall(pattern, html_content)
            for match in matches:
                clean_phone = re.sub(r'[^\d]', '', match)
                if len(clean_phone) >= 9:
                    normalized = self.normalize_phone(clean_phone)
                    if normalized and normalized not in self.owner_phones:
                        return normalized
        
        return None
    
    def extract_statement_uuid_from_url(self, url: str) -> Optional[str]:
        """Extract statement UUID from property URL"""
        match = re.search(r'/pr/(\d+)/', url)
        if match:
            return match.group(1)
        return None
    
    def export_enhanced_excel(self, filename: str = 'owners.xlsx'):
        """Export unique phones to Excel with correct format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            sorted_phones = sorted(self.owner_phones)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Owner Phones"
            
            # Header
            ws['A1'] = "Phone"
            ws['A1'].alignment = Alignment(horizontal='left')
            
            # Phone numbers with correct format
            for i, phone in enumerate(sorted_phones, 2):
                # Format: +995 571 233 844
                digits = re.sub(r'[^\d]', '', phone)
                if len(digits) >= 12:
                    formatted = f"+995 {digits[3:6]} {digits[6:9]} {digits[9:12]}"
                else:
                    formatted = phone
                
                ws[f'A{i}'] = formatted
                ws[f'A{i}'].alignment = Alignment(horizontal='left')
            
            # Save Excel file
            wb.save(filename)
            logger.info(f"✅ Exported {len(sorted_phones)} phones to {filename}")
            return len(sorted_phones)
            
        except Exception as e:
            logger.error(f"Error exporting: {e}")
            return 0
    
    def enhanced_scrape(self):
        """Main enhanced scraping method"""
        logger.info(f"🚀 ENHANCED OWNER SCRAPER STARTING (target: {self.target_phones})")
        
        # Load existing phones to avoid duplicates
        self.load_existing_phones()
        
        if len(self.owner_phones) >= self.target_phones:
            logger.info(f"🎉 Target already reached! {len(self.owner_phones)} phones")
            return len(self.owner_phones)
        
        # Comprehensive URL discovery
        property_urls = self.discover_property_urls_comprehensive(50000)
        logger.info(f"📋 Will scrape {len(property_urls)} properties")
        
        # Scrape phones with progress tracking
        scraped_count = 0
        new_phones_found = 0
        
        for i, url in enumerate(property_urls, 1):
            try:
                phone = self.scrape_property_phone_enhanced(url)
                if phone:
                    old_count = len(self.owner_phones)
                    self.owner_phones.add(phone)
                    if len(self.owner_phones) > old_count:
                        new_phones_found += 1
                        logger.info(f"✅ NEW phone {new_phones_found}: {phone}. Total unique: {len(self.owner_phones)}")
                        
                        # Save progress every 100 new phones
                        if new_phones_found % 100 == 0:
                            self.export_enhanced_excel()
                
                scraped_count += 1
                
                # Progress update
                if i % 500 == 0:
                    success_rate = (new_phones_found / i * 100) if i > 0 else 0
                    logger.info(f"📊 Progress: {i}/{len(property_urls)} URLs. {len(self.owner_phones)} unique phones ({success_rate:.1f}% success)")
                
                # Check if target reached
                if len(self.owner_phones) >= self.target_phones:
                    logger.info(f"🎉 TARGET REACHED! {len(self.owner_phones)} unique phones!")
                    break
                
                time.sleep(0.05)  # Very fast rate limiting
                
            except Exception as e:
                logger.error(f"Error processing URL {i}: {e}")
                continue
        
        # Final export
        final_count = self.export_enhanced_excel()
        
        logger.info(f"🏁 ENHANCED SCRAPING COMPLETE!")
        logger.info(f"   Final unique phones: {final_count}")
        logger.info(f"   Target: {self.target_phones}")
        logger.info(f"   Success: {'YES' if final_count >= self.target_phones else 'PARTIAL'}")
        
        return final_count

def main():
    """Main execution"""
    scraper = EnhancedOwnerScraper()
    result = scraper.enhanced_scrape()
    
    if result >= 10000:
        print(f"🎉 SUCCESS: {result} unique phones collected!")
    else:
        print(f"📊 Collected {result} unique phones (target was 10000)")

if __name__ == "__main__":
    main()
