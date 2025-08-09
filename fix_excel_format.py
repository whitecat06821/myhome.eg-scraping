#!/usr/bin/env python3
"""
Fix Excel CSV format to display phone numbers correctly
Converts existing CSV files to proper text format for Excel
"""

import csv
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def fix_csv_for_excel(input_csv_file, output_excel_file=None):
    """
    Convert CSV file to Excel format with proper text formatting
    """
    if not os.path.exists(input_csv_file):
        logger.error(f"File {input_csv_file} not found")
        return False
    
    if output_excel_file is None:
        # Replace .csv with .xlsx
        output_excel_file = input_csv_file.replace('.csv', '.xlsx')
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Phone Numbers"
        
        # Read the CSV file
        with open(input_csv_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            
            for row_num, row in enumerate(reader, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    if row_num == 1:
                        # Header row
                        cell.value = cell_value
                    else:
                        # Phone number rows - format like client's: +995 571 233 844
                        phone = cell_value.strip()
                        if not phone.startswith('+995'):
                            phone = f"+995{phone}"
                        
                        # Format with spaces: +995 571 233 844
                        if len(phone) >= 13:  # +995XXXXXXXXX
                            spaced_phone = f"{phone[:4]} {phone[4:7]} {phone[7:10]} {phone[10:]}"
                        else:
                            spaced_phone = phone
                        
                        # Set as regular value (General cell type, no apostrophe)
                        cell.value = spaced_phone
        
        # Set column width for better visibility
        ws.column_dimensions['A'].width = 20
        
        # Save the Excel file
        wb.save(output_excel_file)
        logger.info(f"✅ Fixed Excel file saved as: {output_excel_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error converting {input_csv_file}: {e}")
        return False

def fix_csv_with_text_format(input_csv_file, output_csv_file=None):
    """
    Fix CSV file by adding text formatting for Excel
    Uses UTF-8 BOM and special formatting to force Excel text mode
    """
    if not os.path.exists(input_csv_file):
        logger.error(f"File {input_csv_file} not found")
        return False
    
    if output_csv_file is None:
        output_csv_file = input_csv_file.replace('.csv', '_fixed.csv')
    
    try:
        # Read original CSV
        phone_numbers = []
        with open(input_csv_file, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            header = next(reader)  # Skip header
            for row in reader:
                if row and row[0]:
                    phone_numbers.append(row[0])
        
        # Write fixed CSV with UTF-8 BOM and text formatting
        with open(output_csv_file, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow(['Phone'])
            
            # Write phone numbers with proper formatting
            for phone in phone_numbers:
                # Ensure phone starts with +995
                if not phone.startswith('+995'):
                    phone = f"+995{phone.strip()}"
                
                # Add tab character to force text interpretation
                formatted_phone = f"{phone}\t"
                writer.writerow([formatted_phone])
        
        logger.info(f"✅ Fixed CSV file saved as: {output_csv_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error fixing {input_csv_file}: {e}")
        return False

def fix_all_csv_files():
    """Fix all CSV files in current directory"""
    csv_files = [f for f in os.listdir('.') if f.endswith('.csv')]
    
    if not csv_files:
        logger.warning("No CSV files found in current directory")
        return
    
    logger.info(f"Found {len(csv_files)} CSV files to fix: {csv_files}")
    
    for csv_file in csv_files:
        logger.info(f"Processing {csv_file}...")
        
        # Method 1: Create Excel file (.xlsx)
        excel_success = fix_csv_for_excel(csv_file)
        
        # Method 2: Create fixed CSV with text formatting
        csv_success = fix_csv_with_text_format(csv_file)
        
        if excel_success:
            logger.info(f"✅ {csv_file} -> {csv_file.replace('.csv', '.xlsx')} (Excel format)")
        if csv_success:
            logger.info(f"✅ {csv_file} -> {csv_file.replace('.csv', '_fixed.csv')} (Fixed CSV)")

def main():
    """Main function"""
    print("🔧 Excel Format Fixer for Phone Numbers")
    print("=" * 50)
    print("This script will fix CSV files to display phone numbers correctly in Excel")
    print()
    
    # Check if openpyxl is available
    try:
        import openpyxl
        logger.info("✅ openpyxl library available - will create Excel files")
    except ImportError:
        logger.warning("⚠️ openpyxl not available - will only create fixed CSV files")
        logger.info("To install: pip install openpyxl")
    
    # Fix all CSV files
    fix_all_csv_files()
    
    print()
    print("✅ Processing completed!")
    print("📄 Use the .xlsx files for best Excel compatibility")
    print("📄 Or use the _fixed.csv files if you prefer CSV format")

if __name__ == "__main__":
    main()
