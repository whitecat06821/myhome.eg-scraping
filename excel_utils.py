#!/usr/bin/env python3
"""
Excel utilities for proper phone number formatting
"""

import csv
import logging
from typing import List, Set

logger = logging.getLogger(__name__)

def save_phones_to_excel_compatible_csv(phone_numbers: Set[str], filename: str):
    """
    Save phone numbers to CSV format that displays correctly in Excel
    Uses UTF-8 BOM and text formatting to prevent scientific notation
    """
    try:
        # Sort phone numbers
        sorted_phones = sorted(phone_numbers)
        
        # Write CSV with UTF-8 BOM and special formatting
        with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile)
            
            # Write header
            writer.writerow(['Phone'])
            
            # Write phone numbers with proper formatting
            for phone in sorted_phones:
                # Ensure phone starts with +995
                if not phone.startswith('+995'):
                    formatted_phone = f"+995{phone}"
                else:
                    formatted_phone = phone
                
                # Add tab character and space to force text interpretation in Excel
                excel_safe_phone = f" {formatted_phone}"
                writer.writerow([excel_safe_phone])
        
        logger.info(f"Exported {len(sorted_phones)} phone numbers to {filename} (Excel-compatible format)")
        return True
        
    except Exception as e:
        logger.error(f"Error saving to {filename}: {e}")
        return False

def save_phones_to_excel_file(phone_numbers: Set[str], filename: str):
    """
    Save phone numbers directly to Excel file (.xlsx) with client's exact format
    Format: +995 571 233 844 (with spaces, General cell type)
    Requires openpyxl library
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Phone Numbers"
        
        # Set header
        ws['A1'] = 'Phone'
        ws['A1'].font = ws['A1'].font.copy(bold=True)
        
        # Sort phone numbers
        sorted_phones = sorted(phone_numbers)
        
        # Add phone numbers
        for i, phone in enumerate(sorted_phones, 2):
            # Ensure phone starts with +995
            if not phone.startswith('+995'):
                clean_phone = phone.strip()
                formatted_phone = f"+995{clean_phone}"
            else:
                formatted_phone = phone.strip()
            
            # Format with spaces like client's format: +995 571 233 844
            if len(formatted_phone) >= 13:  # +995XXXXXXXXX
                spaced_phone = f"{formatted_phone[:4]} {formatted_phone[4:7]} {formatted_phone[7:10]} {formatted_phone[10:]}"
            else:
                spaced_phone = formatted_phone
            
            # Set as regular value (General cell type, no apostrophe)
            ws[f'A{i}'].value = spaced_phone
            ws[f'A{i}'].alignment = Alignment(horizontal='left')
        
        # Set column width
        ws.column_dimensions['A'].width = 20
        
        # Save Excel file
        wb.save(filename)
        logger.info(f"Exported {len(sorted_phones)} phone numbers to {filename} (Client format: +995 XXX XXX XXX)")
        return True
        
    except ImportError:
        logger.warning("openpyxl not available. Install with: pip install openpyxl")
        return False
    except Exception as e:
        logger.error(f"Error saving Excel file {filename}: {e}")
        return False
