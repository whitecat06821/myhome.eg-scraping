#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ENHANCED Mega Agent Phone Scraper
Advanced discovery methods to reach 10,000+ unique agent phones
"""

import csv
import logging
import time
import re
import json
from typing import Set, List, Dict, Optional
from scraper.fetcher import MyHomeFetcher
from scraper.parser import MyHomeParser

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('enhanced_agent_scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EnhancedAgentScraper:
    """Enhanced agent scraper with multiple discovery methods"""
    
    def __init__(self):
        self.fetcher = MyHomeFetcher()
        self.parser = MyHomeParser()
        self.agent_phones = set()
        self.processed_agents = set()
        self.target_phones = 10000
        self.discovered_agents = set()
        
    def load_existing_agent_phones(self):
        """Load existing agent phones to avoid duplicates"""
        existing_files = [
            'agents.csv',
            'agents.xlsx',
            'master_phones_only.csv'
        ]
        
        for filename in existing_files:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        phone = row.get('Phone', '').strip()
                        if phone:
                            normalized = self.normalize_phone(phone)
                            if normalized:
                                self.agent_phones.add(normalized)
                logger.info(f"Loaded {len(self.agent_phones)} existing agent phones from {filename}")
            except Exception as e:
                logger.debug(f"Could not load {filename}: {e}")
        
        logger.info(f"Starting with {len(self.agent_phones)} existing unique agent phones")
    
    def normalize_phone(self, phone: str) -> str:
        """Normalize phone for deduplication"""
        if not phone:
            return ""
        
        digits = re.sub(r'\D', '', phone)
        
        if len(digits) == 9 and digits.startswith('5'):
            return f"+995{digits}"
        elif len(digits) == 12 and digits.startswith('995'):
            return f"+{digits}"
        elif len(digits) >= 9:
            return f"+995{digits[-9:]}"
        
        return ""
    
    def discover_all_agents_comprehensive(self, max_pages: int = 2000) -> Set[str]:
        """Comprehensive agent discovery using multiple methods"""
        logger.info(f"🔍 COMPREHENSIVE agent discovery (max {max_pages} pages)")
        agent_ids = set()
        
        # Method 1: Main agents API with extensive pagination
        logger.info("📡 Method 1: Main agents API")
        for page in range(1, max_pages + 1):
            try:
                logger.info(f"Fetching agents API page {page}")
                
                api_data = self.fetcher.fetch_agents_api(page)
                if not api_data or not api_data.get('result'):
                    logger.info(f"No data from agents API page {page}")
                    break
                
                agents = self.parser.parse_agents_api_response(api_data)
                
                page_agent_count = 0
                for agent in agents:
                    if 'id' in agent:
                        agent_ids.add(str(agent['id']))
                        page_agent_count += 1
                
                logger.info(f"Page {page}: {page_agent_count} agents. Total: {len(agent_ids)}")
                
                # If no agents found, we've reached the end
                if len(agents) == 0:
                    logger.info(f"No agents found on page {page}. Stopping discovery.")
                    break
                    
                time.sleep(0.2)  # Rate limiting
                
            except Exception as e:
                logger.error(f"Error fetching agents API page {page}: {e}")
                break
        
        # Method 2: Discover through company/broker APIs
        logger.info("📡 Method 2: Company/Broker APIs")
        try:
            company_agents = self.discover_through_company_apis(5000)
            agent_ids.update(company_agents)
            logger.info(f"Added {len(company_agents)} agents through company APIs. Total: {len(agent_ids)}")
        except Exception as e:
            logger.error(f"Error discovering through company APIs: {e}")
        
        # Method 3: Discover through property listings (agents who posted properties)
        logger.info("📡 Method 3: Property listing agents")
        try:
            property_agents = self.discover_through_property_listings(5000)
            agent_ids.update(property_agents)
            logger.info(f"Added {len(property_agents)} agents through property listings. Total: {len(agent_ids)}")
        except Exception as e:
            logger.error(f"Error discovering through property listings: {e}")
        
        # Method 4: Discover through agent detail pages
        logger.info("📡 Method 4: Agent detail page discovery")
        try:
            detail_agents = self.discover_through_agent_details(5000)
            agent_ids.update(detail_agents)
            logger.info(f"Added {len(detail_agents)} agents through detail pages. Total: {len(agent_ids)}")
        except Exception as e:
            logger.error(f"Error discovering through agent details: {e}")
        
        logger.info(f"✅ COMPREHENSIVE agent discovery complete: {len(agent_ids)} unique agent IDs")
        return agent_ids
    
    def discover_through_company_apis(self, max_agents: int = 5000) -> Set[str]:
        """Discover agents through company/broker APIs"""
        company_agents = set()
        
        # Try different company/broker endpoints
        company_endpoints = [
            "/users/company/brokers",
            "/users/company/agents", 
            "/users/brokers",
            "/users/agents"
        ]
        
        for endpoint in company_endpoints:
            try:
                for page in range(1, 100):
                    try:
                        # Try to fetch from company endpoint
                        response = self.fetcher.fetch_page(f"https://api-statements.tnet.ge/v1{endpoint}?page={page}")
                        if not response:
                            break
                        
                        # Try to parse JSON response
                        try:
                            data = json.loads(response)
                            if data.get('result') and data.get('data'):
                                agents = data.get('data', {}).get('data', [])
                                for agent in agents:
                                    if 'id' in agent or 'uuid' in agent:
                                        agent_id = str(agent.get('id') or agent.get('uuid'))
                                        company_agents.add(agent_id)
                                        
                                        if len(company_agents) >= max_agents:
                                            return company_agents
                        except json.JSONDecodeError:
                            # Try to extract agent IDs from HTML
                            agent_matches = re.findall(r'["\']id["\']:\s*(\d+)', response)
                            for match in agent_matches:
                                company_agents.add(match)
                                
                                if len(company_agents) >= max_agents:
                                    return company_agents
                        
                        if len(agents) == 0:
                            break
                            
                        time.sleep(0.1)
                        
                    except Exception as e:
                        logger.debug(f"Error in company endpoint {endpoint} page {page}: {e}")
                        break
                        
            except Exception as e:
                logger.error(f"Error processing company endpoint {endpoint}: {e}")
                continue
        
        return company_agents
    
    def discover_through_property_listings(self, max_agents: int = 5000) -> Set[str]:
        """Discover agents through property listings"""
        property_agents = set()
        
        # Property listing endpoints
        property_endpoints = [
            "/statements",
            "/statements?operation_type_id=1",
            "/statements?operation_type_id=3"
        ]
        
        for endpoint in property_endpoints:
            try:
                for page in range(1, 200):
                    try:
                        api_data = self.fetcher.fetch_property_listings_api(page, endpoint)
                        if not api_data or not api_data.get('result'):
                            break
                        
                        data = api_data.get('data', {})
                        if isinstance(data, dict):
                            items = data.get('data', [])
                        else:
                            items = data
                        
                        if not items:
                            break
                        
                        for item in items:
                            if isinstance(item, dict):
                                # Look for agent information in property data
                                agent_id = item.get('user_id') or item.get('agent_id') or item.get('broker_id')
                                if agent_id:
                                    property_agents.add(str(agent_id))
                                    
                                    if len(property_agents) >= max_agents:
                                        return property_agents
                        
                        time.sleep(0.1)
                        
                    except Exception as e:
                        logger.debug(f"Error in property endpoint {endpoint} page {page}: {e}")
                        break
                        
            except Exception as e:
                logger.error(f"Error processing property endpoint {endpoint}: {e}")
                continue
        
        return property_agents
    
    def discover_through_agent_details(self, max_agents: int = 5000) -> Set[str]:
        """Discover agents through agent detail pages"""
        detail_agents = set()
        
        # Start with some known agent IDs and discover related ones
        known_agent_ids = ["4756", "3777", "1000", "2000", "3000", "4000", "5000"]
        
        for agent_id in known_agent_ids:
            try:
                # Get agent detail
                agent_data = self.fetcher.fetch_agent_detail_api(agent_id)
                if agent_data and agent_data.get('result'):
                    data = agent_data.get('data', {})
                    
                    # Extract related agent IDs
                    related_ids = []
                    
                    # Check for company_id, parent_id, etc.
                    for key in ['company_id', 'parent_id', 'manager_id']:
                        if key in data and data[key]:
                            related_ids.append(str(data[key]))
                    
                    # Check for sub-agents
                    try:
                        sub_agents = self.fetcher.fetch_agent_sub_agents_api(agent_id, 1)
                        if sub_agents and sub_agents.get('result'):
                            sub_data = sub_agents.get('data', {}).get('data', [])
                            for sub_agent in sub_data:
                                if 'id' in sub_agent:
                                    related_ids.append(str(sub_agent['id']))
                    except Exception as e:
                        logger.debug(f"Error getting sub-agents for {agent_id}: {e}")
                    
                    # Add all related IDs
                    for related_id in related_ids:
                        detail_agents.add(related_id)
                        
                        if len(detail_agents) >= max_agents:
                            return detail_agents
                
                time.sleep(0.1)
                
            except Exception as e:
                logger.debug(f"Error processing agent {agent_id}: {e}")
                continue
        
        return detail_agents
    
    def scrape_agent_and_sub_agents_enhanced(self, agent_id: str) -> List[Dict]:
        """Enhanced agent scraping with multiple methods"""
        all_agents = []
        
        try:
            # Method 1: Agent detail API
            agent_data = self.fetcher.fetch_agent_detail_api(agent_id)
            if agent_data and agent_data.get('result'):
                agent_info = agent_data.get('data', {})
                if agent_info:
                    all_agents.append(agent_info)
            
            # Method 2: Sub-agents API
            try:
                sub_agents = self.fetcher.fetch_agent_sub_agents_api(agent_id, 1)
                if sub_agents and sub_agents.get('result'):
                    sub_data = sub_agents.get('data', {}).get('data', [])
                    all_agents.extend(sub_data)
            except Exception as e:
                logger.debug(f"Error getting sub-agents for {agent_id}: {e}")
            
            # Method 3: Try to get company agents if this is a company
            try:
                company_response = self.fetcher.fetch_page(f"https://api-statements.tnet.ge/v1/users/company/{agent_id}/agents")
                if company_response:
                    try:
                        company_data = json.loads(company_response)
                        if company_data.get('result') and company_data.get('data'):
                            company_agents = company_data.get('data', {}).get('data', [])
                            all_agents.extend(company_agents)
                    except json.JSONDecodeError:
                        pass
            except Exception as e:
                logger.debug(f"Error getting company agents for {agent_id}: {e}")
            
        except Exception as e:
            logger.error(f"Error scraping agent {agent_id}: {e}")
        
        return all_agents
    
    def extract_phone_from_agent_enhanced(self, agent: Dict) -> Optional[str]:
        """Enhanced phone extraction from agent data"""
        try:
            # Try multiple phone fields
            phone_fields = ['phone_number', 'phone', 'mobile', 'tel', 'telephone']
            
            for field in phone_fields:
                if field in agent and agent[field]:
                    phone = str(agent[field])
                    normalized = self.normalize_phone(phone)
                    if normalized and normalized not in self.agent_phones:
                        return normalized
            
            # Try to extract from nested data
            if 'contact_info' in agent and isinstance(agent['contact_info'], dict):
                contact = agent['contact_info']
                for field in phone_fields:
                    if field in contact and contact[field]:
                        phone = str(contact[field])
                        normalized = self.normalize_phone(phone)
                        if normalized and normalized not in self.agent_phones:
                            return normalized
            
            return None
            
        except Exception as e:
            logger.debug(f"Error extracting phone from agent: {e}")
            return None
    
    def export_enhanced_agents_excel(self, filename: str = 'agents.xlsx'):
        """Export unique agent phones to Excel with correct format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            sorted_phones = sorted(self.agent_phones)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Agent Phones"
            
            # Header
            ws['A1'] = "Phone"
            ws['A1'].alignment = Alignment(horizontal='left')
            
            # Phone numbers with correct format
            for i, phone in enumerate(sorted_phones, 2):
                # Format: +995 571 233 844
                digits = re.sub(r'[^\d]', '', phone)
                if len(digits) >= 12:
                    formatted = f"+995 {digits[3:6]} {digits[6:9]} {digits[9:12]}"
                else:
                    formatted = phone
                
                ws[f'A{i}'] = formatted
                ws[f'A{i}'].alignment = Alignment(horizontal='left')
            
            # Save Excel file
            wb.save(filename)
            logger.info(f"✅ Exported {len(sorted_phones)} agent phones to {filename}")
            return len(sorted_phones)
            
        except Exception as e:
            logger.error(f"Error exporting agent Excel: {e}")
            return 0
    
    def enhanced_agent_scrape(self):
        """Main enhanced agent scraping method"""
        logger.info(f"🚀 ENHANCED AGENT SCRAPER STARTING (target: {self.target_phones})")
        
        # Load existing agent phones to avoid duplicates
        self.load_existing_agent_phones()
        
        if len(self.agent_phones) >= self.target_phones:
            logger.info(f"🎉 Target already reached! {len(self.agent_phones)} agent phones")
            return len(self.agent_phones)
        
        # Comprehensive agent discovery
        agent_ids = self.discover_all_agents_comprehensive(2000)
        logger.info(f"📋 Will scrape {len(agent_ids)} agents")
        
        # Scrape agent phones with progress tracking
        scraped_count = 0
        new_phones_found = 0
        
        for i, agent_id in enumerate(agent_ids, 1):
            try:
                if agent_id in self.processed_agents:
                    continue
                
                # Scrape agent and sub-agents
                agents = self.scrape_agent_and_sub_agents_enhanced(agent_id)
                
                for agent in agents:
                    phone = self.extract_phone_from_agent_enhanced(agent)
                    if phone:
                        old_count = len(self.agent_phones)
                        self.agent_phones.add(phone)
                        if len(self.agent_phones) > old_count:
                            new_phones_found += 1
                            logger.info(f"✅ NEW agent phone {new_phones_found}: {phone}. Total unique: {len(self.agent_phones)}")
                            
                            # Save progress every 100 new phones
                            if new_phones_found % 100 == 0:
                                self.export_enhanced_agents_excel()
                
                self.processed_agents.add(agent_id)
                scraped_count += 1
                
                # Progress update
                if i % 100 == 0:
                    success_rate = (new_phones_found / i * 100) if i > 0 else 0
                    logger.info(f"📊 Progress: {i}/{len(agent_ids)} agents. {len(self.agent_phones)} unique phones ({success_rate:.1f}% success)")
                
                # Check if target reached
                if len(self.agent_phones) >= self.target_phones:
                    logger.info(f"🎉 TARGET REACHED! {len(self.agent_phones)} unique agent phones!")
                    break
                
                time.sleep(0.1)  # Rate limiting
                
            except Exception as e:
                logger.error(f"Error processing agent {agent_id}: {e}")
                continue
        
        # Final export
        final_count = self.export_enhanced_agents_excel()
        
        logger.info(f"🏁 ENHANCED AGENT SCRAPING COMPLETE!")
        logger.info(f"   Final unique agent phones: {final_count}")
        logger.info(f"   Target: {self.target_phones}")
        logger.info(f"   Success: {'YES' if final_count >= self.target_phones else 'PARTIAL'}")
        
        return final_count

def main():
    """Main execution"""
    scraper = EnhancedAgentScraper()
    result = scraper.enhanced_agent_scrape()
    
    if result >= 10000:
        print(f"🎉 SUCCESS: {result} unique agent phones collected!")
    else:
        print(f"📊 Collected {result} unique agent phones (target was 10000)")

if __name__ == "__main__":
    main()
