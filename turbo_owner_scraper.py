#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Turbo Owner Phone Scraper
Fast parallel scraper to reach 8000+ unique phones quickly
"""

import csv
import logging
import time
import re
# Removed async imports for simplicity
from typing import Set, List, Optional
from urllib.parse import urljoin
import threading
from scraper.fetcher import MyHomeFetcher

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('turbo_scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class TurboOwnerScraper:
    """Fast owner phone scraper with deduplication"""
    
    def __init__(self):
        self.fetcher = MyHomeFetcher()
        self.owner_phones = set()
        self.processed_urls = set()
        self.target_phones = 10000  # Increased target for better coverage
        
    def load_existing_phones(self):
        """Load existing phones to avoid duplicates"""
        existing_files = [
            'agents.csv',
            'owners_direct.csv', 
            'master_phones_only.csv'
        ]
        
        for filename in existing_files:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        phone = row.get('Phone', '').strip()
                        if phone:
                            normalized = self.normalize_phone(phone)
                            if normalized:
                                self.owner_phones.add(normalized)
                logger.info(f"Loaded {len(self.owner_phones)} existing phones from {filename}")
            except Exception as e:
                logger.debug(f"Could not load {filename}: {e}")
        
        logger.info(f"Starting with {len(self.owner_phones)} existing unique phones")
    
    def normalize_phone(self, phone: str) -> str:
        """Normalize phone for deduplication"""
        if not phone:
            return ""
        
        digits = re.sub(r'\D', '', phone)
        
        if len(digits) == 9 and digits.startswith('5'):
            return f"+995{digits}"
        elif len(digits) == 12 and digits.startswith('995'):
            return f"+{digits}"
        elif len(digits) >= 9:
            return f"+995{digits[-9:]}"
        
        return ""
    
    def get_property_urls_fast(self, max_urls: int = 10000) -> List[str]:
        """Fast property URL collection from multiple sources"""
        logger.info(f"🚀 Fast property URL collection (target: {max_urls})")
        
        all_urls = []
        
        # Method 1: Direct API endpoints  
        api_endpoints = [
            "/statements",
            "/statements?operation_type_id=1", 
            "/statements?operation_type_id=3",
        ]
        
        for endpoint in api_endpoints:
            logger.info(f"📡 Fetching from {endpoint}")
            
            for page in range(1, 100):  # Max 100 pages per endpoint
                try:
                    api_data = self.fetcher.fetch_property_listings_api(page, endpoint)
                    if not api_data or not api_data.get('result'):
                        break
                    
                    data = api_data.get('data', {})
                    if isinstance(data, dict):
                        items = data.get('data', [])
                    else:
                        items = data
                    
                    page_urls = []
                    for item in items:
                        if isinstance(item, dict) and 'statement_id' in item:
                            url = f"https://www.myhome.ge/pr/{item['statement_id']}/"
                            page_urls.append(url)
                    
                    all_urls.extend(page_urls)
                    logger.info(f"{endpoint} page {page}: {len(page_urls)} URLs. Total: {len(all_urls)}")
                    
                    if len(page_urls) == 0:
                        break
                    
                    if len(all_urls) >= max_urls:
                        logger.info(f"✅ Reached target of {max_urls} URLs!")
                        return all_urls[:max_urls]
                        
                    time.sleep(0.2)  # Fast rate limiting
                    
                except Exception as e:
                    logger.error(f"Error fetching {endpoint} page {page}: {e}")
                    break
        
        logger.info(f"✅ Collected {len(all_urls)} property URLs")
        return all_urls
    
    def scrape_property_phone(self, property_url: str) -> Optional[str]:
        """Fast phone scraping from property page"""
        try:
            response = self.fetcher.fetch_page(property_url)
            if not response:
                return None
            
            # Fast regex patterns for phone numbers
            phone_patterns = [
                r'\+995\s*\d{3}\s*\d{3}\s*\d{3}',
                r'995\d{9}',
                r'\b5\d{8}\b',
                r'"phone_number":\s*"(\d+)"',  # JSON format
                r'tel:(\d+)',  # Tel links
            ]
            
            for pattern in phone_patterns:
                matches = re.findall(pattern, response)
                for match in matches:
                    clean_phone = re.sub(r'[^\d]', '', match)
                    if len(clean_phone) >= 9:
                        normalized = self.normalize_phone(clean_phone)
                        if normalized and normalized not in self.owner_phones:
                            return normalized
            
            return None
            
        except Exception as e:
            logger.debug(f"Error scraping {property_url}: {e}")
            return None
    
    def export_turbo_excel(self, filename: str = 'owners.xlsx'):
        """Export unique phones to Excel with correct format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            sorted_phones = sorted(self.owner_phones)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Owner Phones"
            
            # Header
            ws['A1'] = "Phone"
            ws['A1'].alignment = Alignment(horizontal='left')
            
            # Phone numbers with correct format
            for i, phone in enumerate(sorted_phones, 2):
                # Format: +995 571 233 844
                digits = re.sub(r'[^\d]', '', phone)
                if len(digits) >= 12:
                    formatted = f"+995 {digits[3:6]} {digits[6:9]} {digits[9:12]}"
                else:
                    formatted = phone
                
                ws[f'A{i}'] = formatted
                ws[f'A{i}'].alignment = Alignment(horizontal='left')
            
            # Save Excel file
            wb.save(filename)
            logger.info(f"✅ Exported {len(sorted_phones)} phones to {filename}")
            return len(sorted_phones)
            
        except Exception as e:
            logger.error(f"Error exporting: {e}")
            return 0
    
    def turbo_scrape(self):
        """Main turbo scraping method"""
        logger.info(f"🚀 TURBO OWNER SCRAPER STARTING (target: {self.target_phones})")
        
        # Load existing phones to avoid duplicates
        self.load_existing_phones()
        
        if len(self.owner_phones) >= self.target_phones:
            logger.info(f"🎉 Target already reached! {len(self.owner_phones)} phones")
            return len(self.owner_phones)
        
        # Get property URLs fast
        property_urls = self.get_property_urls_fast(20000)  # Get many URLs
        logger.info(f"📋 Will scrape {len(property_urls)} properties")
        
        # Scrape phones with progress tracking
        scraped_count = 0
        new_phones_found = 0
        
        for i, url in enumerate(property_urls, 1):
            try:
                phone = self.scrape_property_phone(url)
                if phone:
                    old_count = len(self.owner_phones)
                    self.owner_phones.add(phone)
                    if len(self.owner_phones) > old_count:
                        new_phones_found += 1
                        logger.info(f"✅ NEW phone {new_phones_found}: {phone}. Total unique: {len(self.owner_phones)}")
                        
                        # Save progress every 100 new phones
                        if new_phones_found % 100 == 0:
                            self.export_turbo_excel()
                
                scraped_count += 1
                
                # Progress update
                if i % 500 == 0:
                    success_rate = (new_phones_found / i * 100) if i > 0 else 0
                    logger.info(f"📊 Progress: {i}/{len(property_urls)} URLs. {len(self.owner_phones)} unique phones ({success_rate:.1f}% success)")
                
                # Check if target reached
                if len(self.owner_phones) >= self.target_phones:
                    logger.info(f"🎉 TARGET REACHED! {len(self.owner_phones)} unique phones!")
                    break
                
                time.sleep(0.1)  # Very fast rate limiting
                
            except Exception as e:
                logger.error(f"Error processing URL {i}: {e}")
                continue
        
        # Final export
        final_count = self.export_turbo_excel()
        
        logger.info(f"🏁 TURBO SCRAPING COMPLETE!")
        logger.info(f"   Final unique phones: {final_count}")
        logger.info(f"   Target: {self.target_phones}")
        logger.info(f"   Success: {'YES' if final_count >= self.target_phones else 'PARTIAL'}")
        
        return final_count

def main():
    """Main execution"""
    scraper = TurboOwnerScraper()
    result = scraper.turbo_scrape()
    
    if result >= 10000:
        print(f"🎉 SUCCESS: {result} unique phones collected!")
    else:
        print(f"📊 Collected {result} unique phones (target was 10000)")

if __name__ == "__main__":
    main()
